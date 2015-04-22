import openpyxl
from openpyxl.utils import get_column_letter

# Some config
FILENAME = 'kv008horse.xlsx'
OUTPUT_FILENAME = 'OUTPUT.xlsx'
wb = openpyxl.load_workbook(filename=FILENAME)
sheet_ranges = wb['Лист1']

# Global variables. I know it's stupid but it works.
top_left_corner = None
bottom_right_corner = None
start = None
finish = None


def read():
    global top_left_corner, bottom_right_corner, start, finish

    for y, row in enumerate(sheet_ranges.rows):
        # reading board data
        for x, cell in enumerate(row):
            # finding borders
            if cell.border.top.style is not None and \
                            cell.border.left.style is not None:
                top_left_corner = tuple([x, y])
            # finding obstacles
            if cell.border.bottom.style is not None and \
                            cell.border.right.style is not None:
                bottom_right_corner = tuple([x, y])
            # finding start point
            if cell.value == 's':
                start = tuple([x, y])
            # finding finish point
            if cell.value == 'f':
                finish = tuple([x, y])

    input_matrix = [[True for _ in range(top_left_corner[1], bottom_right_corner[1]+1)]
                    for _ in range(top_left_corner[0], bottom_right_corner[0]+1)]

    # adding obstacles to input matrix
    for y in range(top_left_corner[1], bottom_right_corner[1]+1):
        for x in range(top_left_corner[0], bottom_right_corner[0]+1):
            if sheet_ranges.cell(row=y+1, column=x+1).style.fill.fgColor.rgb == 'FFFF0000':
                input_matrix[x-top_left_corner[0]][y-top_left_corner[1]] = False

    return {'board': input_matrix,
            'start': start,
            'finish': finish,
            'tl_corner': top_left_corner}


def draw(iterable):
    for i in iterable:
        sheet_ranges.cell(row=i.x+top_left_corner[1]+1,
                          column=i.y+top_left_corner[0]+1).value = '█'
    for i in range(len(sheet_ranges.rows[0])):
        sheet_ranges.column_dimensions[get_column_letter(i+1)].width = 3.0
    wb.save(OUTPUT_FILENAME)